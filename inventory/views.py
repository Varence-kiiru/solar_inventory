from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView, FormView, TemplateView
from django.views import View
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.urls import reverse, reverse_lazy, NoReverseMatch
from django.utils.translation import gettext_lazy as _
from django.contrib import messages
from django.conf import settings
from django.core.mail import EmailMessage
from django.db import models, transaction, IntegrityError
from django.db.models import Q, Sum, Count, Avg, F, ExpressionWrapper, DecimalField, Max
from django.db.models.functions import TruncMonth
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic.base import TemplateView, View
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.template.loader import get_template, render_to_string
from datetime import datetime, timedelta
from decimal import Decimal
import os
import csv
import json
import logging
import random
import tempfile
import uuid
import xlsxwriter
from openpyxl import Workbook
from xhtml2pdf import pisa
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from .models import (
    Product, PurchaseOrder, Sale, Customer, Supplier, Location, Brand,
    ProductCategory, SaleItem, PurchaseOrderItem, StockAdjustment, StockMovement,
    HeldSale, HeldSaleItem, ReservedStock, SystemSettings
)
from .forms import (
    ProductForm, PurchaseOrderForm, SaleForm, CustomerForm,
    SupplierForm, BrandForm, ProductCategoryForm, ProductImportForm,
    ProductBulkUpdateForm, PurchaseOrderItemFormSet, PurchaseOrderReceiveForm,
    SaleItemFormSet, StockAdjustmentForm, UserProfileForm, SettingsForm
)

logger = logging.getLogger(__name__)

# Mixin for staff-only views
class StaffRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff

# Custom JSON encoder to handle Decimal objects
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

# Dashboard View
class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get current date and time
        now = timezone.now()
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        thirty_days_ago = now - timedelta(days=30)
        seven_days_ago = now - timedelta(days=7)
        
        # Basic counts
        context['total_products'] = Product.objects.filter(active=True).count()
        context['low_stock_count'] = Product.objects.filter(
            quantity_in_stock__lte=F('minimum_stock_level'),
            active=True
        ).count()
        context['out_of_stock_count'] = Product.objects.filter(
            quantity_in_stock=0,
            active=True
        ).count()
        context['in_stock_count'] = context['total_products'] - context['low_stock_count']
        
        # Monthly sales
        monthly_sales = Sale.objects.filter(
            sale_date__gte=start_of_month,
            status='completed'
        ).aggregate(total=Sum('total_amount'))
        context['monthly_sales'] = monthly_sales['total'] or 0
        
        # Pending orders
        context['pending_orders'] = PurchaseOrder.objects.filter(
            status='pending'
        ).count()
        
        # Recent sales
        context['recent_sales'] = Sale.objects.filter(
            status__in=['completed', 'proforma']
        ).order_by('-sale_date')[:5]
        
        # Low stock products
        context['low_stock_products'] = Product.objects.filter(
            quantity_in_stock__lte=F('minimum_stock_level'),
            active=True
        ).order_by('quantity_in_stock')[:5]
        
        # Recent purchase orders
        context['recent_purchase_orders'] = PurchaseOrder.objects.all().order_by('-order_date')[:5]
        
        # Sales chart data (last 7 days)
        sales_data = []
        labels = []
        
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
            
            daily_sales = Sale.objects.filter(
                sale_date__range=(day_start, day_end),
                status='completed'
            ).aggregate(total=Sum('total_amount'))
            
            # Convert Decimal to float for JSON serialization
            total = daily_sales['total'] or 0
            if isinstance(total, Decimal):
                total = float(total)
                
            sales_data.append(total)
            labels.append(day.strftime('%d %b'))
        
        # Use the custom encoder for JSON serialization
        context['sales_chart_data'] = json.dumps(sales_data, cls=DecimalEncoder)
        context['sales_chart_labels'] = json.dumps(labels)
        
        # Top selling products
        top_products = SaleItem.objects.filter(
            sale__status='completed',
            sale__sale_date__gte=thirty_days_ago
        ).values('product__name').annotate(
            total_sold=Sum('quantity')
        ).order_by('-total_sold')[:5]
        
        if top_products:
            context['top_products'] = True
            top_products_labels = [item['product__name'] for item in top_products]
            
            # Convert Decimal to float for JSON serialization
            top_products_data = []
            for item in top_products:
                total_sold = item['total_sold']
                if isinstance(total_sold, Decimal):
                    total_sold = float(total_sold)
                top_products_data.append(total_sold)
            
            context['top_products_labels'] = json.dumps(top_products_labels)
            context['top_products_data'] = json.dumps(top_products_data, cls=DecimalEncoder)
        else:
            context['top_products'] = False
        
        # System status metrics
        # These are example calculations - adjust based on your business logic
        
        # Inventory health (percentage of products not in low stock)
        if context['total_products'] > 0:
            context['inventory_health'] = round(
                ((context['total_products'] - context['low_stock_count']) / context['total_products']) * 100
            )
        else:
            context['inventory_health'] = 100
            
        # Set color based on health
        if context['inventory_health'] >= 70:
            context['inventory_health_color'] = 'success'
        elif context['inventory_health'] >= 40:
            context['inventory_health_color'] = 'warning'
        else:
            context['inventory_health_color'] = 'danger'
        
        # Pending orders percentage
        total_orders = PurchaseOrder.objects.count()
        if total_orders > 0:
            context['pending_orders_percentage'] = round(
                (context['pending_orders'] / total_orders) * 100
            )
        else:
            context['pending_orders_percentage'] = 0
        
        # Sales target (example - assuming a monthly target of 100,000)
        monthly_target = 100000
        monthly_sales_value = float(context['monthly_sales']) if isinstance(context['monthly_sales'], Decimal) else context['monthly_sales']
        context['sales_target_percentage'] = min(
            round((monthly_sales_value / monthly_target) * 100), 100
        )
        
        # Customer satisfaction (example - could be based on actual feedback)
        # For demo purposes, we'll use a random value between 70-100
        context['customer_satisfaction'] = random.randint(80, 100)
        
        # Last update time
        context['last_update'] = now
        
        return context

# Brand Views
class BrandListView(LoginRequiredMixin, ListView):
    model = Brand
    template_name = 'inventory/brands/list.html'
    context_object_name = 'brand_list'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        
        if search:
            queryset = queryset.filter(name__icontains=search)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        return context

class BrandCreateView(LoginRequiredMixin, CreateView):
    model = Brand
    form_class = BrandForm
    template_name = 'inventory/brands/form.html'
    success_url = reverse_lazy('inventory:brand-list')
    
    def form_valid(self, form):
        messages.success(self.request, "Brand created successfully!")
        return super().form_valid(form)

class BrandUpdateView(LoginRequiredMixin, UpdateView):
    model = Brand
    form_class = BrandForm
    template_name = 'inventory/brands/form.html'
    success_url = reverse_lazy('inventory:brand-list')
    
    def form_valid(self, form):
        messages.success(self.request, "Brand updated successfully!")
        return super().form_valid(form)

class BrandDeleteView(LoginRequiredMixin, DeleteView):
    model = Brand
    template_name = 'inventory/brands/confirm_delete.html'
    success_url = reverse_lazy('inventory:brand-list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Brand deleted successfully!")
        return super().delete(request, *args, **kwargs)

class BrandDetailView(LoginRequiredMixin, DetailView):
    model = Brand
    template_name = 'inventory/brands/detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        brand = self.get_object()
        
        # Get products for this brand
        context['products'] = Product.objects.filter(brand=brand)
        
        return context

# Product Views
class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'inventory/products/list.html'
    context_object_name = 'products'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        category = self.request.GET.get('category')
        brand = self.request.GET.get('brand')
        stock_status = self.request.GET.get('stock_status')
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(description__icontains=search)
            )

        if category:
            queryset = queryset.filter(category_id=category)

        if brand:
            queryset = queryset.filter(brand_id=brand)
            
        if stock_status == 'low':
            queryset = Product.get_low_stock_products()
        elif stock_status == 'out':
            queryset = queryset.filter(quantity_in_stock=0)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = ProductCategory.objects.all()
        context['brands'] = Brand.objects.all()
        context['needs_restock'] = Product.get_low_stock_products()

        # Get filter parameters for maintaining state
        context['search'] = self.request.GET.get('search', '')
        context['selected_category'] = self.request.GET.get('category', '')
        context['selected_brand'] = self.request.GET.get('brand', '')
        context['selected_stock_status'] = self.request.GET.get('stock_status', '')

        return context

class ProductExportView(LoginRequiredMixin, View):
    def get(self, request):
        export_format = request.GET.get('format', 'csv')
        products = Product.objects.all()
        
        # Apply filters if any
        if 'search' in request.GET and request.GET['search']:
            search = request.GET['search']
            products = products.filter(
                Q(name__icontains=search) | 
                Q(sku__icontains=search) |
                Q(barcode__icontains=search)
            )

        if 'category' in request.GET and request.GET['category']:
            products = products.filter(category_id=request.GET['category'])

        if 'brand' in request.GET and request.GET['brand']:
            products = products.filter(brand_id=request.GET['brand'])

        if 'supplier' in request.GET and request.GET['supplier']:
            products = products.filter(suppliers__id=request.GET['supplier'])

        if 'in_stock' in request.GET:
            products = products.filter(quantity_in_stock__gt=0)
            
        if 'low_stock' in request.GET:
            products = products.filter(quantity_in_stock__lte=F('min_stock_level'))

        if 'active' in request.GET:
            products = products.filter(active=True)

        if export_format == 'csv':
            return self._export_csv(products)
        elif export_format == 'excel':
            return self._export_excel(products)
        elif export_format == 'pdf':
            return self._export_pdf(products)
        else:
            return HttpResponse("Unsupported export format", status=400)
    
    def _export_csv(self, products):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products.csv"'

        writer = csv.writer(response)
        writer.writerow(['Name', 'SKU', 'Category', 'Brand', 'Price', 'Cost', 'Stock', 'Status'])

        for product in products:
            writer.writerow([
                product.name,
                product.sku,
                product.category.name if product.category else '',
                product.brand.name if product.brand else '',
                product.unit_price,
                product.cost_price,
                product.quantity_in_stock,
                'Active' if product.active else 'Inactive'
            ])

        return response
    
    def _export_excel(self, products):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Add headers
        headers = ['Name', 'SKU', 'Category', 'Brand', 'Price', 'Cost', 'Stock', 'Status']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Add data
        for row_num, product in enumerate(products, 2):
            ws.cell(row=row_num, column=1, value=product.name)
            ws.cell(row=row_num, column=2, value=product.sku)
            ws.cell(row=row_num, column=3, value=product.category.name if product.category else '')
            ws.cell(row=row_num, column=4, value=product.brand.name if product.brand else '')
            ws.cell(row=row_num, column=5, value=product.unit_price)
            ws.cell(row=row_num, column=6, value=product.cost_price)
            ws.cell(row=row_num, column=7, value=product.quantity_in_stock)
            ws.cell(row=row_num, column=8, value='Active' if product.active else 'Inactive')
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        
        return response

    def _export_pdf(self, products):
        # This is a simplified version - you might want to use a proper PDF library
        # like ReportLab, WeasyPrint, or xhtml2pdf for a real implementation
        
        # For now, we'll just return a simple text response
        response = HttpResponse(
            "PDF export not implemented yet. Please use CSV or Excel export.",
            content_type='text/plain'
        )
        response['Content-Disposition'] = 'attachment; filename="products.txt"'

        return response

class ProductExportTemplateView(LoginRequiredMixin, View):
    """View for downloading an empty product import template"""
    
    def get(self, request):
        export_format = request.GET.get('format', 'csv')
        
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="product_import_template.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Name', 'SKU', 'Category', 'Brand', 'Unit Price', 'Cost Price', 
                            'Quantity', 'Min Stock Level', 'Description', 'Active'])
            # Add a sample row
            writer.writerow(['Sample Product', 'SKU123', 'Solar Panels', 'SunPower', '199.99', 
                            '149.99', '10', '5', 'Sample product description', 'Yes'])
            
            return response
            
        elif export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Products Template"
            
            # Add headers
            headers = ['Name', 'SKU', 'Category', 'Brand', 'Unit Price', 'Cost Price', 
                      'Quantity', 'Min Stock Level', 'Description', 'Active']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Add sample row
            sample_data = ['Sample Product', 'SKU123', 'Solar Panels', 'SunPower', 199.99, 
                          149.99, 10, 5, 'Sample product description', 'Yes']
            for col_num, value in enumerate(sample_data, 1):
                ws.cell(row=2, column=col_num, value=value)
            
            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Create response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="product_import_template.xlsx"'
            
            return response
        
        else:
            return HttpResponse("Unsupported format", status=400)


class ProductImportView(LoginRequiredMixin, FormView):
    """View for importing products from CSV or Excel files"""
    
    template_name = 'inventory/products/import.html'
    form_class = ProductImportForm
    success_url = reverse_lazy('inventory:product-import-result')
    
    def form_valid(self, form):
        file = form.cleaned_data['file']
        update_existing = form.cleaned_data['update_existing']
        skip_errors = form.cleaned_data['skip_errors']
        
        # Store import options in session for the result view
        self.request.session['import_options'] = {
            'update_existing': update_existing,
            'skip_errors': skip_errors,
        }
        
        # Process the file
        results = self.process_import_file(file, update_existing, skip_errors)
        
        # Store results in session for the result view
        self.request.session['import_results'] = results
        
        return super().form_valid(form)
    
    def process_import_file(self, file, update_existing, skip_errors):
        """Process the imported file and return results"""
        results = {
            'success': True,
            'message': 'Import completed successfully',
            'created_count': 0,
            'updated_count': 0,
            'error_count': 0,
            'errors': [],
            'created_products': [],
            'updated_products': []
        }
        
        # Determine file type and process accordingly
        if file.name.endswith('.csv'):
            # Process CSV file
            try:
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                
                for row_num, row in enumerate(reader, start=2):  # Start at 2 to account for header row
                    try:
                        self.process_row(row, update_existing, results)
                    except Exception as e:
                        results['error_count'] += 1
                        results['errors'].append({
                            'row': row_num,
                            'identifier': row.get('SKU', row.get('Name', f'Row {row_num}')),
                            'message': str(e)
                        })
                        if not skip_errors:
                            results['success'] = False
                            results['message'] = f'Import failed at row {row_num}'
                            break
            
            except Exception as e:
                results['success'] = False
                results['message'] = f'Error processing CSV file: {str(e)}'
        
        elif file.name.endswith(('.xlsx', '.xls')):
            # Process Excel file
            try:
                wb = load_workbook(filename=BytesIO(file.read()))
                ws = wb.active
                
                # Get headers from first row
                headers = [cell.value for cell in ws[1]]
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Convert row to dict using headers
                        row_dict = {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
                        self.process_row(row_dict, update_existing, results)
                    except Exception as e:
                        results['error_count'] += 1
                        sku = row[headers.index('SKU')] if 'SKU' in headers else None
                        name = row[headers.index('Name')] if 'Name' in headers else None
                        identifier = sku or name or f'Row {row_num}'
                        
                        results['errors'].append({
                            'row': row_num,
                            'identifier': identifier,
                            'message': str(e)
                        })
                        if not skip_errors:
                            results['success'] = False
                            results['message'] = f'Import failed at row {row_num}'
                            break
            
            except Exception as e:
                results['success'] = False
                results['message'] = f'Error processing Excel file: {str(e)}'
        
        else:
            results['success'] = False
            results['message'] = 'Unsupported file format. Please upload a CSV or Excel file.'
        
        return results
    
    def process_row(self, row, update_existing, results):
        """Process a single row from the import file"""
        # Clean and validate data
        sku = row.get('SKU', '').strip()
        name = row.get('Name', '').strip()
        
        if not sku or not name:
            raise ValueError("SKU and Name are required fields")
        
        # Check if product exists
        try:
            product = Product.objects.get(sku=sku)
            if update_existing:
                # Update existing product
                self.update_product(product, row)
                results['updated_count'] += 1
                results['updated_products'].append(product)
            else:
                raise ValueError(f"Product with SKU '{sku}' already exists")
        except Product.DoesNotExist:
            # Create new product
            product = self.create_product(row)
            results['created_count'] += 1
            results['created_products'].append(product)
    
    def create_product(self, row):
        """Create a new product from import data"""
        # Get or create category if provided
        category = None
        if row.get('Category'):
            category, _ = ProductCategory.objects.get_or_create(
                name=row['Category'].strip()
            )
        
        # Get or create brand if provided
        brand = None
        if row.get('Brand'):
            brand, _ = Brand.objects.get_or_create(
                name=row['Brand'].strip()
            )
        
        # Create the product
        product = Product(
            name=row['Name'].strip(),
            sku=row['SKU'].strip(),
            description=row.get('Description', '').strip(),
            category=category,
            brand=brand,
            unit_price=Decimal(str(row.get('Unit Price', 0))),
            cost_price=Decimal(str(row.get('Cost Price', 0))),
            quantity_in_stock=int(float(row.get('Quantity', 0))),
            minimum_stock_level=int(float(row.get('Min Stock Level', 0))),
            active=self.parse_boolean(row.get('Active', 'Yes'))
        )
        product.save()
        return product
    
    def update_product(self, product, row):
        """Update an existing product from import data"""
        # Update category if provided
        if row.get('Category'):
            category, _ = ProductCategory.objects.get_or_create(
                name=row['Category'].strip()
            )
            product.category = category
        
        # Update brand if provided
        if row.get('Brand'):
            brand, _ = Brand.objects.get_or_create(
                name=row['Brand'].strip()
            )
            product.brand = brand
        
        # Update other fields if provided
        if row.get('Name'):
            product.name = row['Name'].strip()
        
        if row.get('Description'):
            product.description = row['Description'].strip()
        
        if row.get('Unit Price'):
            product.unit_price = Decimal(str(row['Unit Price']))
        
        if row.get('Cost Price'):
            product.cost_price = Decimal(str(row['Cost Price']))
        
        if row.get('Quantity'):
            product.quantity_in_stock = int(float(row['Quantity']))
        
        if row.get('Min Stock Level'):
            product.minimum_stock_level = int(float(row['Min Stock Level']))
        
        if row.get('Active'):
            product.active = self.parse_boolean(row['Active'])
        
        product.save()
        return product
    
    def parse_boolean(self, value):
        """Parse boolean values from various formats"""
        if isinstance(value, bool):
            return value
        
        if isinstance(value, str):
            return value.lower() in ('yes', 'true', '1', 'y', 't')
        
        return bool(value)


class ProductImportResultView(LoginRequiredMixin, TemplateView):
    """View for displaying the results of a product import"""
    
    template_name = 'inventory/products/import_result.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get import results from session
        results = self.request.session.get('import_results', {})
        
        # Add results to context
        context.update(results)
        
        # Clear session data
        if 'import_results' in self.request.session:
            del self.request.session['import_results']
        
        if 'import_options' in self.request.session:
            del self.request.session['import_options']
        
        return context


class ProductBulkUpdateView(LoginRequiredMixin, FormView):
    """View for bulk updating multiple products"""
    
    template_name = 'inventory/products/bulk_update.html'
    form_class = ProductBulkUpdateForm
    success_url = reverse_lazy('inventory:product-list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Pass empty initial data if not POST
        if self.request.method != 'POST':
            kwargs['initial'] = {}
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get product IDs from request
        product_ids = self.request.GET.getlist('ids') or self.request.POST.getlist('product_ids')
        products = Product.objects.filter(id__in=product_ids)
        
        context['products'] = products
        context['product_ids'] = product_ids
        
        return context
    
    def form_valid(self, form):
        # Get product IDs and clean data
        product_ids = self.request.POST.getlist('product_ids')
        products = Product.objects.filter(id__in=product_ids)
        cleaned_data = form.cleaned_data
        
        # Track which fields to update
        fields_to_update = {}
        for field_name, value in cleaned_data.items():
            if value is not None:  # Only update fields that were included in the form
                fields_to_update[field_name] = value
        
        # Update products
        update_count = 0
        for product in products:
            updated = False
            for field_name, value in fields_to_update.items():
                setattr(product, field_name, value)
                updated = True
            
            if updated:
                product.save()
                update_count += 1
        
        messages.success(self.request, f"{update_count} products updated successfully.")
        return super().form_valid(form)


class ProductBulkDeleteView(LoginRequiredMixin, View):
    """View for bulk deleting multiple products"""
    
    def post(self, request):
        product_ids = request.POST.getlist('product_ids')
        
        if not product_ids:
            messages.error(request, "No products selected for deletion.")
            return redirect('inventory:product-list')
        
        # Delete products
        deleted_count = Product.objects.filter(id__in=product_ids).delete()[0]
        
        messages.success(request, f"{deleted_count} products deleted successfully.")
        return redirect('inventory:product-list')

class ProductDetailView(LoginRequiredMixin, DetailView):
    model = Product
    template_name = 'inventory/products/detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.get_object()

        # Get sales history for this product
        context['sales_history'] = SaleItem.objects.filter(
            product=product
        ).select_related('sale').order_by('-sale__sale_date')[:10]

        # Get purchase history for this product
        context['purchase_history'] = PurchaseOrderItem.objects.filter(
            product=product
        ).select_related('purchase_order').order_by('-purchase_order__order_date')[:10]

        return context

class ProductCreateView(LoginRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/products/form.html'
    success_url = reverse_lazy('inventory:product-list')

    def form_valid(self, form):
        messages.success(self.request, "Product created successfully!")
        return super().form_valid(form)

class ProductRestockView(LoginRequiredMixin, View):
    """View for restocking a product."""
    
    def get(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        suppliers = Supplier.objects.all()
        
        context = {
            'product': product,
            'suppliers': suppliers,
            'title': _('Restock Product')
        }
        
        return render(request, 'inventory/products/product_restock.html', context)
    
    def post(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        quantity = int(request.POST.get('quantity', 0))
        supplier_id = request.POST.get('supplier')
        cost_price = Decimal(request.POST.get('cost_price', 0))
        notes = request.POST.get('notes', '')
       
        if quantity <= 0:
            messages.error(request, _("Quantity must be greater than zero."))
            return redirect('inventory:product-detail', pk=pk)
           
        # Update product stock
        product.quantity_in_stock += quantity
        product.last_restocked = timezone.now()
       
        # Update cost price if it changed
        if cost_price > 0 and cost_price != product.cost_price:
            product.cost_price = cost_price
           
        product.save()
       
        # Create stock movement record
        StockMovement.objects.create(
            product=product,
            quantity=quantity,
            movement_type='restock',
            notes=notes,
            created_by=request.user
        )
       
        messages.success(request, _(f"Successfully added {quantity} units to inventory."))
        return redirect('inventory:product-detail', pk=pk)

class ProductAdjustStockView(LoginRequiredMixin, View):
    """View for adjusting product stock."""
    
    def post(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        adjustment_type = request.POST.get('adjustment_type')
        quantity = int(request.POST.get('quantity', 0))
        reason = request.POST.get('reason')
        notes = request.POST.get('notes', '')
        
        if quantity < 0:
            messages.error(request, _("Quantity cannot be negative."))
            return redirect('inventory:product-detail', pk=pk)
            
        original_quantity = product.quantity_in_stock
        
        # Adjust stock based on adjustment type
        if adjustment_type == 'add':
            product.quantity_in_stock += quantity
            movement_type = 'stock_add'
            movement_quantity = quantity
            message = _(f"Successfully added {quantity} units to inventory.")
        elif adjustment_type == 'remove':
            if quantity > product.quantity_in_stock:
                messages.error(request, _("Cannot remove more than available stock."))
                return redirect('inventory:product-detail', pk=pk)
            product.quantity_in_stock -= quantity
            movement_type = 'stock_remove'
            movement_quantity = -quantity
            message = _(f"Successfully removed {quantity} units from inventory.")
        elif adjustment_type == 'set':
            movement_quantity = quantity - product.quantity_in_stock
            product.quantity_in_stock = quantity
            movement_type = 'stock_set'
            message = _(f"Stock level set to {quantity} units.")
        
        product.save()
        
        # Create stock movement record
        StockMovement.objects.create(
            product=product,
            quantity=movement_quantity,
            movement_type=movement_type,
            notes=f"Reason: {reason}. {notes}",
            created_by=request.user
        )
        
        messages.success(request, message)
        return redirect('inventory:product-detail', pk=pk)

class ProductUpdateView(LoginRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/products/form.html'
    success_url = reverse_lazy('inventory:product-list')

    def form_valid(self, form):
        messages.success(self.request, "Product updated successfully!")
        return super().form_valid(form)

class ProductDeleteView(LoginRequiredMixin, DeleteView):
    model = Product
    template_name = 'inventory/products/confirm_delete.html'
    success_url = reverse_lazy('inventory:product-list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Product deleted successfully!")
        return super().delete(request, *args, **kwargs)

# Category Views
class CategoryListView(LoginRequiredMixin, ListView):
    model = ProductCategory
    template_name = 'inventory/categories/list.html'
    context_object_name = 'category_list'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')

        if search:
            queryset = queryset.filter(name__icontains=search)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['search'] = self.request.GET.get('search', '')

        return context

class CategoryCreateView(LoginRequiredMixin, CreateView):
    model = ProductCategory
    form_class = ProductCategoryForm
    template_name = 'inventory/categories/form.html'
    success_url = reverse_lazy('inventory:category-list')

    def form_valid(self, form):
        messages.success(self.request, "Category created successfully!")
        return super().form_valid(form)

class CategoryUpdateView(LoginRequiredMixin, UpdateView):
    model = ProductCategory
    form_class = ProductCategoryForm
    template_name = 'inventory/categories/form.html'
    success_url = reverse_lazy('inventory:category-list')
    
    def form_valid(self, form):
        messages.success(self.request, "Category updated successfully!")
        return super().form_valid(form)

class CategoryDeleteView(LoginRequiredMixin, DeleteView):
    model = ProductCategory
    template_name = 'inventory/categories/confirm_delete.html'
    success_url = reverse_lazy('inventory:category-list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Category deleted successfully!")
        return super().delete(request, *args, **kwargs)

# Customer Views
class CustomerListView(LoginRequiredMixin, ListView):
    model = Customer
    template_name = 'inventory/customers/list.html'
    context_object_name = 'customers'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        return context

class CustomerDetailView(LoginRequiredMixin, DetailView):
    model = Customer
    template_name = 'inventory/customers/detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = self.get_object()
        
        # Get sales history for this customer
        context['sales_history'] = Sale.objects.filter(
            customer=customer
        ).order_by('-sale_date')
        
        # Calculate customer statistics
        total_spent = Sale.objects.filter(
            customer=customer
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        order_count = Sale.objects.filter(customer=customer).count()
        
        if order_count > 0:
            average_order = total_spent / order_count
        else:
            average_order = 0
            
        context['total_spent'] = total_spent
        context['order_count'] = order_count
        context['average_order'] = average_order
        
        return context

class CustomerCreateView(LoginRequiredMixin, CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'inventory/customers/form.html'
    success_url = reverse_lazy('inventory:customer-list')

    def form_valid(self, form):
        messages.success(self.request, "Customer created successfully!")
        return super().form_valid(form)

class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'inventory/customers/form.html'
    success_url = reverse_lazy('inventory:customer-list')

    def form_valid(self, form):
        messages.success(self.request, "Customer updated successfully!")
        return super().form_valid(form)

class CustomerDeleteView(LoginRequiredMixin, DeleteView):
    model = Customer
    template_name = 'inventory/customers/confirm_delete.html'
    success_url = reverse_lazy('inventory:customer-list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Customer deleted successfully!")
        return super().delete(request, *args, **kwargs)

# Supplier Views
class SupplierListView(LoginRequiredMixin, ListView):
    model = Supplier
    template_name = 'inventory/suppliers/list.html'
    context_object_name = 'suppliers'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Apply search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(contact_person__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        
        # Apply status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        # Apply sorting
        sort = self.request.GET.get('sort', 'name')
        if sort == 'name':
            queryset = queryset.order_by('name')
        elif sort == 'email':
            queryset = queryset.order_by('email')
        elif sort == 'recent':
            queryset = queryset.order_by('-created_at')
        
        # Annotate with product count - FIXED: use a different name for the annotation
        queryset = queryset.annotate(products_count=Count('products'))
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['status'] = self.request.GET.get('status', '')
        context['sort'] = self.request.GET.get('sort', 'name')
        return context

class SupplierDetailView(LoginRequiredMixin, DetailView):
    model = Supplier
    template_name = 'inventory/suppliers/detail.html'
    context_object_name = 'supplier'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        supplier = self.get_object()
        
        # Get products from this supplier - FIXED: use the correct relationship
        context['supplier_products'] = supplier.products.all()
        
        # Get recent purchase orders
        context['recent_purchase_orders'] = supplier.purchase_orders.all().order_by('-order_date')[:5]
        
        return context

class SupplierCreateView(LoginRequiredMixin, CreateView):
    model = Supplier
    form_class = SupplierForm
    template_name = 'inventory/suppliers/form.html'
    success_url = reverse_lazy('inventory:supplier-list')

    def form_valid(self, form):
        messages.success(self.request, _("Supplier created successfully!"))
        return super().form_valid(form)

    def form_invalid(self, form):
        # Add this to debug form validation errors
        messages.error(self.request, f"Form validation failed: {form.errors}")
        return super().form_invalid(form)

class SupplierUpdateView(LoginRequiredMixin, UpdateView):
    model = Supplier
    form_class = SupplierForm
    template_name = 'inventory/suppliers/form.html'

    def get_success_url(self):
        return reverse_lazy('inventory:supplier-detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, _("Supplier updated successfully!"))
        return super().form_valid(form)

class SupplierDeleteView(LoginRequiredMixin, DeleteView):
    model = Supplier
    template_name = 'inventory/suppliers/confirm_delete.html'
    success_url = reverse_lazy('inventory:supplier-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        supplier = self.get_object()

        # Count related records - FIXED: use the correct relationship names
        context['related_products'] = supplier.products.count()
        context['related_orders'] = supplier.purchase_orders.count()

        return context

    def delete(self, request, *args, **kwargs):
        supplier = self.get_object()
        try:
            result = super().delete(request, *args, **kwargs)
            messages.success(request, _("Supplier '{}' deleted successfully!").format(supplier.name))
            return result
        except ProtectedError:
            messages.error(
                request, 
                _("Cannot delete supplier '{}' because it has related records that would be affected.").format(supplier.name)
            )
            return redirect('inventory:supplier-detail', pk=supplier.pk)



# Purchase Order Views
class PurchaseOrderListView(LoginRequiredMixin, ListView):
    model = PurchaseOrder
    template_name = 'inventory/purchase_orders/list.html'
    context_object_name = 'orders'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        supplier = self.request.GET.get('supplier')

        if search:
            queryset = queryset.filter(
                Q(order_number__icontains=search) |
                Q(supplier__name__icontains=search)
            )
            
        if status:
            queryset = queryset.filter(status=status)
            
        if supplier:
            queryset = queryset.filter(supplier_id=supplier)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['suppliers'] = Supplier.objects.all()
        context['search'] = self.request.GET.get('search', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_supplier'] = self.request.GET.get('supplier', '')
        
        # Status choices for filtering
        context['status_choices'] = PurchaseOrder.STATUS_CHOICES
        
        return context

class PurchaseOrderDetailView(LoginRequiredMixin, DetailView):
    model = PurchaseOrder
    template_name = 'inventory/purchase_orders/detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        purchase_order = self.get_object()
        
        # Get items for this purchase order
        context['items'] = PurchaseOrderItem.objects.filter(
            purchase_order=purchase_order
        ).select_related('product')
        
        return context

class PurchaseOrderCreateView(LoginRequiredMixin, CreateView):
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = 'inventory/purchase_orders/form.html'
    success_url = reverse_lazy('inventory:purchase-order-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = PurchaseOrderItemFormSet(self.request.POST)
        else:
            context['formset'] = PurchaseOrderItemFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            
            formset.instance = self.object
            formset.save()
            
            # Calculate and save total amount
            subtotal = sum(item.quantity * item.unit_cost for item in self.object.items.all())
            self.object.total_amount = subtotal + self.object.shipping_cost + self.object.tax_amount
            self.object.save()
            
            messages.success(self.request, _("Purchase order created successfully!"))
            return redirect(self.success_url)
        return self.render_to_response(self.get_context_data(form=form))

class PurchaseOrderUpdateView(LoginRequiredMixin, UpdateView):
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = 'inventory/purchase_orders/form.html'
    success_url = reverse_lazy('inventory:purchase-order-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = PurchaseOrderItemFormSet(
                self.request.POST, instance=self.object
            )
        else:
            context['formset'] = PurchaseOrderItemFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            
            # Calculate and save total amount
            subtotal = sum(item.quantity * item.unit_cost for item in self.object.items.all())
            self.object.total_amount = subtotal + self.object.shipping_cost + self.object.tax_amount
            self.object.save()
            
            messages.success(self.request, _("Purchase order updated successfully!"))
            return redirect(self.success_url)
        return self.render_to_response(self.get_context_data(form=form))

class PurchaseOrderDeleteView(LoginRequiredMixin, DeleteView):
    model = PurchaseOrder
    template_name = 'inventory/purchase_orders/confirm_delete.html'
    success_url = reverse_lazy('inventory:purchase-order-list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Purchase order deleted successfully!")
        return super().delete(request, *args, **kwargs)

class PurchaseOrderReceiveView(LoginRequiredMixin, UpdateView):
    model = PurchaseOrder
    template_name = 'inventory/purchase_orders/receive.html'
    form_class = PurchaseOrderReceiveForm
    success_url = reverse_lazy('inventory:purchase-order-list')
   
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        purchase_order = self.get_object()
       
        # Get items for this purchase order
        context['items'] = PurchaseOrderItem.objects.filter(
            purchase_order=purchase_order
        ).select_related('product')
        
        # Add today's date for the received_date field
        context['today'] = timezone.now()
       
        return context
   
    def post(self, request, *args, **kwargs):
        purchase_order = self.get_object()
        
        # Get received date
        received_date = request.POST.get('received_date')
        notes = request.POST.get('notes', '')
        
        if not received_date:
            messages.error(request, _("Received date is required"))
            return self.get(request, *args, **kwargs)
        
        try:
            received_date = datetime.strptime(received_date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, _("Invalid date format"))
            return self.get(request, *args, **kwargs)
        
        # Process each item
        items = PurchaseOrderItem.objects.filter(purchase_order=purchase_order)
        all_received = True
        
        for item in items:
            quantity_key = f'item_{item.id}_quantity'
            notes_key = f'item_{item.id}_notes'
            
            try:
                receive_quantity = int(request.POST.get(quantity_key, 0))
            except ValueError:
                receive_quantity = 0
                
            item_notes = request.POST.get(notes_key, '')
            
            # Validate quantity
            remaining = item.quantity - item.received_quantity
            if receive_quantity < 0 or receive_quantity > remaining:
                messages.error(request, _(f"Invalid quantity for {item.product.name}"))
                return self.get(request, *args, **kwargs)
            
            if receive_quantity > 0:
                # Update received quantity
                item.received_quantity += receive_quantity
                item.notes = item_notes if item_notes else item.notes
                item.save()
                
                # Update product inventory
                product = item.product
                product.quantity_in_stock += receive_quantity
                product.save()
            
            # Check if all items are fully received
            if item.received_quantity < item.quantity:
                all_received = False
        
        # Update purchase order status
        if all_received:
            purchase_order.status = PurchaseOrder.RECEIVED
        elif any(item.received_quantity > 0 for item in items):
            purchase_order.status = PurchaseOrder.PARTIALLY_RECEIVED
        
        purchase_order.received_date = received_date
        purchase_order.received_by = request.user
        purchase_order.notes = notes if notes else purchase_order.notes
        purchase_order.save()
        
        messages.success(request, _("Items received successfully!"))
        return redirect('inventory:purchase-order-detail', pk=purchase_order.id)

# Stock Adjustment Views
class StockAdjustmentListView(LoginRequiredMixin, ListView):
    model = StockAdjustment
    template_name = 'inventory/stock_adjustments/list.html'
    context_object_name = 'adjustments'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        
        if search:
            queryset = queryset.filter(
                Q(reference_number__icontains=search) |
                Q(product__name__icontains=search) |
                Q(reason__icontains=search)
            )
        
        return queryset.order_by('-date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        return context

class StockAdjustmentCreateView(LoginRequiredMixin, CreateView):
    model = StockAdjustment
    form_class = StockAdjustmentForm
    template_name = 'inventory/stock_adjustments/form.html'
    success_url = reverse_lazy('inventory:stock-adjustment-list')

    def form_valid(self, form):
        adjustment = form.save(commit=False)

        # Set the adjusted_by field to the current user
        adjustment.adjusted_by = self.request.user

        # Let the model's save method handle the product quantity update
        adjustment.save()

        messages.success(self.request, "Stock adjustment recorded successfully!")
        return super().form_valid(form)

class StockAdjustmentDetailView(LoginRequiredMixin, DetailView):
    model = StockAdjustment
    template_name = 'inventory/stock_adjustments/detail.html'
    context_object_name = 'adjustment'

class StockAdjustmentUpdateView(LoginRequiredMixin, UpdateView):
    model = StockAdjustment
    form_class = StockAdjustmentForm
    template_name = 'inventory/stock_adjustments/form.html'
    success_url = reverse_lazy('inventory:stock-adjustment-list')
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Optionally disable certain fields that shouldn't be changed after creation
        # For example, you might not want to allow changing the product or adjustment type
        # form.fields['product'].disabled = True
        # form.fields['adjustment_type'].disabled = True
        return form
    
    def form_valid(self, form):
        # Get the original adjustment before it's updated
        original_adjustment = self.get_object()
        original_product = original_adjustment.product
        original_quantity = original_adjustment.quantity
        original_type = original_adjustment.adjustment_type
        
        # Save the form but don't commit to DB yet
        adjustment = form.save(commit=False)
        
        # If product, quantity, or adjustment type has changed, we need to:
        # 1. Reverse the original adjustment's effect on stock
        # 2. Apply the new adjustment's effect
        if (original_product != adjustment.product or 
            original_quantity != adjustment.quantity or 
            original_type != adjustment.adjustment_type):
            
            # Reverse original adjustment
            if original_type == StockAdjustment.INCREASE:
                original_product.quantity_in_stock -= original_quantity
            else:  # DECREASE
                original_product.quantity_in_stock += original_quantity
            original_product.save()
            
            # Apply new adjustment
            new_product = adjustment.product
            if adjustment.adjustment_type == StockAdjustment.INCREASE:
                new_product.quantity_in_stock += adjustment.quantity
            else:  # DECREASE
                # Ensure we don't go below zero
                new_quantity = max(0, new_product.quantity_in_stock - adjustment.quantity)
                new_product.quantity_in_stock = new_quantity
            new_product.save()
        
        # Now save the adjustment
        adjustment.save()
        
        messages.success(self.request, "Stock adjustment updated successfully!")
        return super().form_valid(form)

class StockAdjustmentDeleteView(LoginRequiredMixin, DeleteView):
    model = StockAdjustment
    template_name = 'inventory/stock_adjustments/confirm_delete.html'
    success_url = reverse_lazy('inventory:stock-adjustment-list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Stock adjustment deleted successfully!")
        return super().delete(request, *args, **kwargs)

# Stock Count View
class StockCountView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/stock_count/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.all().order_by('name')
        return context

    def post(self, request, *args, **kwargs):
        product_ids = request.POST.getlist('product_id')
        counted_quantities = request.POST.getlist('counted_quantity')
        
        for i in range(len(product_ids)):
            product_id = product_ids[i]
            counted_quantity = int(counted_quantities[i])
            
            product = Product.objects.get(id=product_id)
            current_quantity = product.quantity_in_stock
            
            if counted_quantity != current_quantity:
                # Create stock adjustment
                adjustment = StockAdjustment()
                adjustment.product = product
                adjustment.reason = "Stock count adjustment"
                adjustment.adjusted_by = request.user  # Add this line to set the adjusted_by field
                
                if counted_quantity > current_quantity:
                    adjustment.adjustment_type = 'increase'
                    adjustment.quantity = counted_quantity - current_quantity
                else:
                    adjustment.adjustment_type = 'decrease'
                    adjustment.quantity = current_quantity - counted_quantity
                
                adjustment.save()
                
                # Update product quantity
                product.quantity_in_stock = counted_quantity
                product.save()
        
        messages.success(request, "Stock count completed successfully!")
        return redirect('inventory:stock-adjustment-list')

# Sales Views
class SaleListView(LoginRequiredMixin, ListView):
    model = Sale
    template_name = 'inventory/sales/list.html'
    context_object_name = 'sales'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        customer = self.request.GET.get('customer')

        if search:
            queryset = queryset.filter(
                Q(invoice_number__icontains=search) |
                Q(customer__name__icontains=search)
            )

        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d')
                queryset = queryset.filter(sale_date__gte=date_from)
            except ValueError:
                pass
                
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d')
                date_to = date_to.replace(hour=23, minute=59, second=59)
                queryset = queryset.filter(sale_date__lte=date_to)
            except ValueError:
                pass
                
        if customer:
            queryset = queryset.filter(customer_id=customer)

        return queryset.order_by('-sale_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['customers'] = Customer.objects.all()
        context['search'] = self.request.GET.get('search', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['selected_customer'] = self.request.GET.get('customer', '')

        return context

class SaleCreateView(LoginRequiredMixin, CreateView):
    model = Sale
    form_class = SaleForm
    template_name = 'inventory/sales/sale_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['items_formset'] = SaleItemFormSet(self.request.POST)
        else:
            context['items_formset'] = SaleItemFormSet()
        return context

    def form_valid(self, form):
        form.instance.sales_person = self.request.user
        context = self.get_context_data()
        items_formset = context['items_formset']

        if items_formset.is_valid():
            try:
                with transaction.atomic():
                    self.object = form.save(commit=False)
                    
                    # Generate a unique invoice number if not provided
                    if not self.object.invoice_number:
                        last_sale = Sale.objects.order_by('-id').first()
                        if last_sale and last_sale.invoice_number:
                            try:
                                # Try to extract the numeric part and increment
                                last_num = int(last_sale.invoice_number.split('-')[-1])
                                self.object.invoice_number = f"INV-{last_num + 1:06d}"
                            except (ValueError, IndexError):
                                # Fallback if parsing fails
                                self.object.invoice_number = f"INV-{timezone.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6]}"
                        else:
                            self.object.invoice_number = f"INV-000001"

                    # Configure credit sales and proforma invoices
                    if self.object.payment_method == Sale.CREDIT or self.object.is_proforma:
                        self.object.paid = False
                        self.object.is_proforma = True

                        if not self.object.customer:
                            return JsonResponse({'success': False, 'error': _('Credit sales require a customer')})

                        if not self.object.due_date:
                            self.object.due_date = timezone.now().date() + timedelta(days=30)

                    self.object.save()
                    items_formset.instance = self.object
                    items_formset.save()

                    # Create stock adjustments for all sales (both proforma and regular)
                    for item in self.object.items.all():
                        product = item.product
                       
                        # Check if there's enough stock
                        if item.quantity > product.quantity_in_stock:
                            return JsonResponse({
                                'success': False,
                                'error': _('Not enough stock for {}. Only {} available.').format(
                                    product.name, product.quantity_in_stock
                                )
                            })
                       
                        # For proforma sales, create a reserved stock record instead of reducing inventory
                        if self.object.is_proforma:
                            # Create or update reserved stock record
                            reserved_stock, created = ReservedStock.objects.get_or_create(
                                product=product,
                                sale=self.object,
                                defaults={'quantity': item.quantity}
                            )
                            if not created:
                                reserved_stock.quantity += item.quantity
                                reserved_stock.save()
                               
                            # Log the reservation
                            logger.info(f"Reserved {item.quantity} units of {product.name} for proforma invoice #{self.object.invoice_number}")
                        else:
                            # For regular sales, reduce inventory directly
                            product.quantity_in_stock -= item.quantity
                            product.save()
                           
                            # Create stock movement record
                            StockMovement.objects.create(
                                product=product,
                                movement_type='sale',
                                quantity=item.quantity,
                                reference_id=self.object.id,  # This is now a valid field
                                notes=f"Sale #{self.object.invoice_number}",
                                created_by=self.request.user
                            )

                    if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                        redirect_url = reverse('inventory:sale-detail', kwargs={'pk': self.object.id})
                        return JsonResponse({
                            'success': True,
                            'redirect_url': redirect_url,
                            'is_proforma': self.object.is_proforma
                        })

                    messages.success(self.request, _('Sale completed successfully'))
                    return redirect(self.object.get_absolute_url())

            except Exception as e:
                logger.error(f"Error processing sale: {str(e)}")
                return JsonResponse({'success': False, 'error': str(e)})

        return self.form_invalid(form)

@require_POST
def proforma_convert(request):
    """Convert a proforma invoice to a regular sale"""
    proforma_id = request.POST.get('proforma_id')
    payment_method = request.POST.get('payment_method')
    reference_number = request.POST.get('reference_number')
    amount_tendered = request.POST.get('amount_tendered')
    change_amount = request.POST.get('change_amount')
    notes = request.POST.get('notes')

    try:
        # Get the proforma invoice
        proforma = Sale.objects.get(id=proforma_id, is_proforma=True)

        # Validate payment method
        if payment_method not in dict(Sale.PAYMENT_METHOD_CHOICES):
            return JsonResponse({'success': False, 'error': _('Invalid payment method')})

        # Validate reference number for card/mobile payments
        if payment_method in [Sale.CARD, Sale.MOBILE] and not reference_number:
            return JsonResponse({'success': False, 'error': _('Reference number is required for card/mobile payments')})

        # Validate amount tendered for cash payments
        if payment_method == Sale.CASH and Decimal(amount_tendered or 0) < proforma.total_amount:
            return JsonResponse({'success': False, 'error': _('Amount tendered must be at least equal to the total amount')})

        # Convert the proforma to a sale
        try:
            proforma.convert_to_sale(
                payment_method=payment_method,
                reference_number=reference_number,
                amount_tendered=Decimal(amount_tendered or 0),
                change_amount=Decimal(change_amount or 0)
            )
            
            # Update notes if provided
            if notes:
                proforma.notes = notes
                proforma.save()
                
            return JsonResponse({
                'success': True,
                'redirect_url': reverse('inventory:sale-detail', kwargs={'pk': proforma.id})
            })
        except ValidationError as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
    except Sale.DoesNotExist:
        return JsonResponse({'success': False, 'error': _('Proforma invoice not found')}, status=404)
    except Exception as e:
        logger.error(f"Error converting proforma invoice: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def mark_sale_paid(request, pk):
    if request.method == 'POST':
        try:
            sale = Sale.objects.get(pk=pk)
            
            # Update paid status
            sale.paid = True
            
            # If this was a proforma invoice, it's now a regular paid invoice
            if sale.is_proforma:
                sale.is_proforma = False
                
                # If the invoice number starts with PRO-, change it to INV-
                if sale.invoice_number.startswith('PRO-'):
                    # Generate a new invoice number
                    last_sale = Sale.objects.filter(
                        invoice_number__startswith='INV-'
                    ).order_by('-invoice_number').first()
                    
                    if last_sale:
                        try:
                            last_num = int(last_sale.invoice_number.split('-')[-1])
                            new_invoice_num = f"INV-{last_num + 1:06d}"
                        except (ValueError, IndexError):
                            new_invoice_num = f"INV-{timezone.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6]}"
                    else:
                        new_invoice_num = "INV-00001"
                    
                    sale.invoice_number = new_invoice_num
            
            sale.save()
            
            # Handle any reserved stock for this sale
            reserved_stocks = ReservedStock.objects.filter(sale=sale)
            for reserved in reserved_stocks:
                product = reserved.product
                
                # Deduct the reserved quantity from actual stock
                if reserved.quantity > product.quantity_in_stock:
                    messages.error(
                        request, 
                        _(f'Not enough stock for {product.name}. Only {product.quantity_in_stock} available.')
                    )
                    return redirect('inventory:invoice-list')
                
                product.quantity_in_stock -= reserved.quantity
                product.save()
                
                # Create stock movement record
                StockMovement.objects.create(
                    product=product,
                    movement_type='sale',
                    quantity=reserved.quantity,
                    reference_id=sale.id,
                    notes=f"Sale #{sale.invoice_number} (converted from proforma)",
                    created_by=request.user
                )
                
                # Delete the reserved stock record
                reserved.delete()
            
            messages.success(request, _('Invoice marked as paid successfully.'))
            
            # Determine where to redirect based on the referrer
            referer = request.META.get('HTTP_REFERER', '')
            if 'invoice' in referer:
                return redirect('inventory:invoice-list')
            else:
                return redirect('inventory:sale-list')
                
        except Sale.DoesNotExist:
            messages.error(request, _('Invoice not found.'))
        except Exception as e:
            messages.error(request, str(e))
    
    # Default redirect to invoice list
    return redirect('inventory:invoice-list')

class SaleDetailView(LoginRequiredMixin, DetailView):
    model = Sale
    template_name = 'inventory/sales/detail.html'
    context_object_name = 'sale'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sale = self.get_object()

        # Get items for this sale
        context['items'] = SaleItem.objects.filter(
            sale=sale
        ).select_related('product')

        context['is_proforma'] = self.object.is_proforma
        return context

class POSView(LoginRequiredMixin, CreateView):
    model = Sale
    form_class = SaleForm
    template_name = 'inventory/sales/pos.html'
    success_url = reverse_lazy('inventory:pos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.filter(quantity_in_stock__gt=0).order_by('name')
        context['categories'] = ProductCategory.objects.all()
        context['customers'] = Customer.objects.filter(is_active=True).order_by('name')
        context['customer_form'] = CustomerForm()
        context['currency_symbol'] = 'KES'
        context['default_tax_rate'] = 16

        # Pre-select customer if provided in URL
        customer_id = self.request.GET.get('customer')
        if customer_id:
            try:
                context['selected_customer'] = Customer.objects.get(id=customer_id)
            except Customer.DoesNotExist:
                pass

        return context

    def form_invalid(self, form):
        # Add this to debug form validation errors
        messages.error(self.request, f"Form validation failed: {form.errors}")
        return super().form_invalid(form)

    def form_valid(self, form):
        try:
            with transaction.atomic():  # Use transaction to ensure data integrity
                sale = form.save(commit=False)

                # Set the sales person to the current user
                sale.sales_person = self.request.user

                # Generate invoice number
                last_sale = Sale.objects.order_by('-id').first()
                if last_sale:
                    last_id = last_sale.id
                else:
                    last_id = 0

                # Use different prefix for proforma invoices
                if sale.is_proforma:
                    sale.invoice_number = f"PRO-{last_id + 1:06d}"
                else:
                    sale.invoice_number = f"INV-{last_id + 1:06d}"

                # Process sale items
                items_data_str = self.request.POST.get('items_data', '[]')
                try:
                    items_data = json.loads(items_data_str)
                except json.JSONDecodeError:
                    messages.error(self.request, f"Invalid items data format: {items_data_str}")
                    return self.form_invalid(form)

                if not items_data:
                    messages.error(self.request, "No items added to the sale.")
                    return self.form_invalid(form)

                # First check if all products have enough stock
                for item_data in items_data:
                    product_id = item_data.get('product_id')
                    quantity = int(item_data.get('quantity', 0))

                    if not product_id or quantity <= 0:
                        messages.error(self.request, "Invalid product or quantity.")
                        return self.form_invalid(form)
                    
                    try:
                        product = Product.objects.get(id=product_id)
                        if product.quantity_in_stock < quantity:
                            messages.error(
                                self.request,
                                f"Not enough stock for {product.name}. Available: {product.quantity_in_stock}"
                            )
                            return self.form_invalid(form)
                    except Product.DoesNotExist:
                        messages.error(self.request, f"Product with ID {product_id} does not exist.")
                        return self.form_invalid(form)
            
                # Save the sale first
                sale.save()
            
                # Now process the items
                total_amount = 0
                for item_data in items_data:
                    product = Product.objects.get(id=item_data['product_id'])
                    quantity = int(item_data['quantity'])
                    price = float(item_data['price'])

                    # Create sale item
                    sale_item = SaleItem(
                        sale=sale,
                        product=product,
                        quantity=quantity,
                        unit_price=price
                    )
                    sale_item.save()

                    # Handle inventory based on sale type
                    if sale.is_proforma:
                        # For proforma invoices, create reserved stock record
                        reserved_stock, created = ReservedStock.objects.get_or_create(
                            product=product,
                            sale=sale,
                            defaults={'quantity': quantity}
                        )
                        if not created:
                            reserved_stock.quantity += quantity
                            reserved_stock.save()
                        
                        # Create stock movement record for reservation
                        StockMovement.objects.create(
                            product=product,
                            movement_type='reservation',
                            quantity=quantity,
                            reference_id=sale.id,
                            notes=f"Reserved for Proforma #{sale.invoice_number}"
                        )
                        
                        # Log the reservation
                        logger.info(f"Reserved {quantity} units of {product.name} for proforma invoice #{sale.invoice_number}")
                    else:
                        # For regular sales, reduce inventory
                        product.quantity_in_stock -= quantity
                        product.save()
                        
                        # Create stock movement record for sale
                        StockMovement.objects.create(
                            product=product,
                            movement_type='sale',
                            quantity=quantity,
                            reference_id=sale.id,
                            notes=f"Sale #{sale.invoice_number}"
                        )

                    # Add to total
                    total_amount += quantity * price

                # Update sale total
                sale.total_amount = total_amount
                sale.save()

                # Set appropriate success message
                if sale.is_proforma:
                    messages.success(self.request, "Proforma invoice created successfully!")
                else:
                    messages.success(self.request, "Sale completed successfully!")

                # Redirect to invoice view
                return redirect('inventory:sale-detail', pk=sale.id)

        except Exception as e:
            # Log the full exception for debugging
            logger.error(f"Error processing sale: {str(e)}", exc_info=True)
            messages.error(self.request, f"Error processing sale: {str(e)}")
            return self.form_invalid(form)

# Invoice List View
class InvoiceListView(LoginRequiredMixin, ListView):
    model = Sale
    template_name = 'inventory/invoices/list.html'
    context_object_name = 'invoices'
    paginate_by = 20

    def get_queryset(self):
        return Sale.objects.all().order_by('-sale_date')

# View for displaying the invoice in the browser with buttons
class InvoiceView(LoginRequiredMixin, DetailView):
    model = Sale
    template_name = 'inventory/invoices/invoice_view.html'
    context_object_name = 'sale'

    def get_context_data(self, **kwargs):
        """Add additional context data to the template."""
        context = super().get_context_data(**kwargs)
        return context

    def post(self, request, *args, **kwargs):
        """Handle POST requests to mark the sale as paid."""
        self.object = self.get_object()
        if self.object.paid:
            messages.warning(request, _('This sale is already marked as paid.'))
            return self.get(request, *args, **kwargs)

        try:
            with transaction.atomic():
                self.object.paid = True
                if self.object.is_proforma:
                    self.object.is_proforma = False
                    self.object.status = Sale.STATUS_COMPLETED
                    # Generate a new unique invoice number
                    self.object.invoice_number = self.generate_new_invoice_number()
                    # Update inventory for all items in the sale
                    for item in self.object.items.all():
                        item.product.quantity_in_stock -= item.quantity
                        item.product.save()
                else:
                    self.object.status = Sale.STATUS_COMPLETED

                self.object.save()
                messages.success(request, _('Sale has been marked as paid.'))
        except Exception as e:
            logger.error(f"Error marking sale as paid: {str(e)}")
            messages.error(request, _('An error occurred while marking the sale as paid.'))

        return redirect('inventory:sale-detail', pk=self.object.pk)

    def generate_new_invoice_number(self):
        """Generate a new unique invoice number."""
        prefix = "INV"
        max_id = 0
        highest_invoice = Sale.objects.filter(
            invoice_number__startswith=f"{prefix}-"
        ).exclude(pk=self.object.pk).order_by('-invoice_number').first()

        if highest_invoice:
            try:
                max_id = int(highest_invoice.invoice_number.split('-')[1])
            except (IndexError, ValueError):
                max_id = self.object.pk

        new_invoice_number = f"{prefix}-{max_id + 1:06d}"

        while Sale.objects.filter(invoice_number=new_invoice_number).exists():
            max_id += 1
            new_invoice_number = f"{prefix}-{max_id + 1:06d}"

        return new_invoice_number

# View for generating PDF in the browser
class InvoicePDFView(LoginRequiredMixin, DetailView):
    model = Sale

    def get(self, request, *args, **kwargs):
        sale = self.get_object()
        template = get_template('inventory/invoices/invoice_pdf.html')

        # Get the absolute path to the logo
        logo_path = self.get_logo_path()

        # Render the template with the logo path
        html = template.render({
            'sale': sale,
            'logo_path': logo_path
        })

        # Create a PDF file
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

        # Return the PDF for viewing in the browser
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="Invoice_{sale.invoice_number}.pdf"'
            return response

        return HttpResponse("Error generating PDF", status=400)

    def get_logo_path(self):
        """Get the path to the company logo."""
        logo_path = os.path.join(settings.STATIC_ROOT, 'img/company-logo.png')

        if not os.path.exists(logo_path):
            for static_dir in settings.STATICFILES_DIRS:
                potential_path = os.path.join(static_dir, 'img/company-logo.png')
                if os.path.exists(potential_path):
                    logo_path = potential_path
                    break

        return logo_path if os.path.exists(logo_path) else None

def invoice_pdf_download(request, pk):
    """Generate a downloadable PDF invoice"""
    sale = get_object_or_404(Sale, pk=pk)
    template = get_template('inventory/invoices/invoice_pdf.html')

    # Use a properly formatted static URL for the logo
    logo_url = request.build_absolute_uri(settings.STATIC_URL + 'img/company-logo.png')

    # Render the template with the logo URL
    html = template.render({
        'sale': sale,
        'logo_url': logo_url  # Pass the correct logo URL
    })

    # Create a PDF file
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    # Return the PDF as a downloadable file
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f"Invoice_{sale.invoice_number}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return HttpResponse("Error generating PDF", status=400)

def download_invoice(request):
    """Download an invoice as PDF"""
    sale_id = request.GET.get('sale_id')

    try:
        sale = Sale.objects.get(id=sale_id)
        
        # Generate PDF invoice
        pdf_content = generate_invoice_pdf(sale)
        
        # Create response with PDF content
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{sale.invoice_number}.pdf"'
        
        return response
    except Sale.DoesNotExist:
        messages.error(request, 'Sale not found')
        return redirect('inventory:pos')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('inventory:pos')

# Helper function to generate PDF invoice
def generate_invoice_pdf(sale, is_proforma=False):
    """Generate a PDF invoice for a sale"""
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # Set up document properties
    title = "Proforma Invoice" if is_proforma else "Invoice"
    p.setTitle(f"{title} #{sale.invoice_number}")
    
    # Add company logo and header
    # p.drawImage("path/to/logo.png", 50, 700, width=100, height=50)
    
    # Add invoice title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 750, f"{title} #{sale.invoice_number}")
    
    # Add date
    p.setFont("Helvetica", 10)
    p.drawString(50, 730, f"Date: {sale.sale_date.strftime('%d/%m/%Y')}")
    
    if sale.due_date:
        p.drawString(50, 715, f"Due Date: {sale.due_date.strftime('%d/%m/%Y')}")
    
    # Add customer info
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, 680, "Customer Information:")
    p.setFont("Helvetica", 10)
    
    if sale.customer:
        p.drawString(50, 665, f"Name: {sale.customer.name}")
        if sale.customer.phone:
            p.drawString(50, 650, f"Phone: {sale.customer.phone}")
        if sale.customer.email:
            p.drawString(50, 635, f"Email: {sale.customer.email}")
    else:
        p.drawString(50, 665, "Walk-in Customer")
    
    # Add items table header
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, 600, "Product")
    p.drawString(250, 600, "Quantity")
    p.drawString(320, 600, "Unit Price")
    p.drawString(400, 600, "Total")
    
    # Draw a line under the header
    p.line(50, 595, 550, 595)
    
    # Add items
    y_position = 580
    p.setFont("Helvetica", 10)
    
    for item in sale.items.all():
        p.drawString(50, y_position, item.product.name[:30])
        p.drawString(250, y_position, str(item.quantity))
        p.drawString(320, y_position, f"Ksh. {item.unit_price}")
        p.drawString(400, y_position, f"Ksh. {item.total_price}")
        y_position -= 15
        
        # Check if we need a new page
        if y_position < 100:
            p.showPage()
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, 750, "Product")
            p.drawString(250, 750, "Quantity")
            p.drawString(320, 750, "Unit Price")
            p.drawString(400, 750, "Total")
            p.line(50, 745, 550, 745)
            p.setFont("Helvetica", 10)
            y_position = 730
    
    # Draw a line above the summary
    p.line(50, y_position - 5, 550, y_position - 5)
    
    # Add summary
    y_position -= 20
    p.drawString(320, y_position, "Subtotal:")
    p.drawString(400, y_position, f"Ksh. {sale.subtotal}")
    
    if sale.discount_percentage > 0:
        y_position -= 15
        p.drawString(320, y_position, f"Discount ({sale.discount_percentage}%):")
        p.drawString(400, y_position, f"Ksh. {sale.discount_amount}")
    
    if sale.tax_percentage > 0:
        y_position -= 15
        p.drawString(320, y_position, f"Tax ({sale.tax_percentage}%):")
        p.drawString(400, y_position, f"Ksh. {sale.tax_amount}")
    
    if sale.shipping_cost:
        y_position -= 15
        p.drawString(320, y_position, "Shipping:")
        p.drawString(400, y_position, f"Ksh. {sale.shipping_cost}")
    
    # Total
    y_position -= 20
    p.setFont("Helvetica-Bold", 10)
    p.drawString(320, y_position, "Total:")
    p.drawString(400, y_position, f"Ksh. {sale.total_amount}")
    
    # Add payment status
    y_position -= 30
    p.setFont("Helvetica", 10)
    status = "PAID" if sale.paid else "UNPAID"
    p.drawString(50, y_position, f"Payment Status: {status}")
    
    if sale.payment_method:
        p.drawString(50, y_position - 15, f"Payment Method: {sale.get_payment_method_display()}")
    
    # Add notes if any
    if sale.notes:
        y_position -= 40
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y_position, "Notes:")
        p.setFont("Helvetica", 10)
        
        # Split notes into multiple lines if needed
        notes_lines = [sale.notes[i:i+80] for i in range(0, len(sale.notes), 80)]
        for line in notes_lines:
            y_position -= 15
            p.drawString(50, y_position, line)
    
    # Add footer
    p.setFont("Helvetica", 8)
    p.drawString(50, 50, "Thank you for your business!")
    p.drawString(50, 35, "For questions about this invoice, please contact us.")
    
    # Close the PDF object cleanly
    p.showPage()
    p.save()
    
    # Get the value of the BytesIO buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content

# Report Views
class SalesReportView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/reports/sales.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Date range filter
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        sales_query = Sale.objects.all()

        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d')
                sales_query = sales_query.filter(sale_date__gte=date_from)
            except ValueError:
                pass

        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d')
                date_to = date_to.replace(hour=23, minute=59, second=59)
                sales_query = sales_query.filter(sale_date__lte=date_to)
            except ValueError:
                pass

        # Total sales
        total_sales = sales_query.aggregate(
            total=Sum('total_amount')
        )['total'] or 0

        # Sales by product
        sales_by_product = SaleItem.objects.filter(
            sale__in=sales_query
        ).values(
            'product__name'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_amount=Sum(F('quantity') * F('unit_price'))
        ).order_by('-total_amount')

        # Sales by customer
        sales_by_customer = sales_query.values(
            'customer__name'
        ).annotate(
            total_sales=Count('id'),
            total_amount=Sum('total_amount')
        ).order_by('-total_amount')

        # Sales by month
        sales_by_month = sales_query.annotate(
            month=TruncMonth('sale_date')
        ).values('month').annotate(
            total=Sum('total_amount')
        ).order_by('month')

        context['total_sales'] = total_sales
        context['sales_by_product'] = sales_by_product
        context['sales_by_customer'] = sales_by_customer
        context['sales_by_month'] = sales_by_month
        context['date_from'] = date_from
        context['date_to'] = date_to

        return context
    
    def post(self, request, *args, **kwargs):
        # Export sales report as CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sales_report.csv"'

        # Date range filter
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')

        sales_query = Sale.objects.all()

        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d')
                sales_query = sales_query.filter(sale_date__gte=date_from)
            except ValueError:
                pass
                
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d')
                date_to = date_to.replace(hour=23, minute=59, second=59)
                sales_query = sales_query.filter(sale_date__lte=date_to)
            except ValueError:
                pass
        
        writer = csv.writer(response)
        writer.writerow(['Invoice', 'Date', 'Customer', 'Total Amount'])

        for sale in sales_query:
            writer.writerow([
                sale.invoice_number,
                sale.sale_date.strftime('%Y-%m-%d'),
                sale.customer.name if sale.customer else 'Walk-in Customer',
                sale.total_amount
            ])
            
        return response

class InventoryReportView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/reports/inventory.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Filter by category
        category = self.request.GET.get('category')

        products_query = Product.objects.all()

        if category:
            products_query = products_query.filter(category_id=category)
        
        # Low stock products
        context['low_stock'] = Product.get_low_stock_products()
        
        # Out of stock products
        context['out_of_stock'] = products_query.filter(quantity_in_stock=0)
        
        # Inventory value
        inventory_value = products_query.aggregate(
            total=Sum(F('quantity_in_stock') * F('cost_price'))
        )['total'] or 0
        
        # Products by category
        products_by_category = products_query.values(
            'category__name'
        ).annotate(
            count=Count('id'),
            total_value=Sum(F('quantity_in_stock') * F('cost_price'))
        ).order_by('-total_value')
        
        # Top products by value
        top_products = products_query.annotate(
            inventory_value=ExpressionWrapper(
                F('quantity_in_stock') * F('cost_price'),
                output_field=DecimalField()
            )
        ).order_by('-inventory_value')[:10]
        
        context['inventory_value'] = inventory_value
        context['products_by_category'] = products_by_category
        context['top_products'] = top_products
        context['categories'] = ProductCategory.objects.all()
        context['selected_category'] = category
        
        return context
    
    def post(self, request, *args, **kwargs):
        # Export inventory report as CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'
        
        # Filter by category
        category = request.POST.get('category')
        
        products_query = Product.objects.all()
        
        if category:
            products_query = products_query.filter(category_id=category)
        
        writer = csv.writer(response)
        writer.writerow(['SKU', 'Name', 'Category', 'Quantity', 'Cost Price', 'Value'])
        
        for product in products_query:
            writer.writerow([
                product.sku,
                product.name,
                product.category.name if product.category else 'Uncategorized',
                product.quantity_in_stock,
                product.cost_price,
                product.quantity_in_stock * product.cost_price
            ])
            
        return response

class CustomerReportView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/reports/customers.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Top customers by sales
        top_customers = Customer.objects.annotate(
            total_sales=Count('sales'),
            total_spent=Sum('sales__total_amount')
        ).filter(total_sales__gt=0).order_by('-total_spent')[:10]
        
        # Inactive customers (no purchases in last 90 days)
        ninety_days_ago = timezone.now() - timedelta(days=90)
        inactive_customers = Customer.objects.annotate(
            last_purchase=Max('sales__sale_date')
        ).filter(
            Q(last_purchase__lt=ninety_days_ago) | Q(last_purchase__isnull=True)
        )
        
        # New customers (last 30 days)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        new_customers = Customer.objects.filter(
            created_at__gte=thirty_days_ago
        )

        context['top_customers'] = top_customers
        context['inactive_customers'] = inactive_customers
        context['new_customers'] = new_customers
        
        return context
    
    def post(self, request, *args, **kwargs):
        # Export customer report as CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="customer_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Name', 'Email', 'Phone', 'Total Orders', 'Total Spent', 'Last Purchase'])
        
        customers = Customer.objects.annotate(
            total_orders=Count('sales'),
            total_spent=Sum('sales__total_amount'),
            last_purchase=Max('sales__sale_date')
        )
        
        for customer in customers:
            writer.writerow([
                customer.name,
                customer.email,
                customer.phone,
                customer.total_orders,
                customer.total_spent or 0,
                customer.last_purchase.strftime('%Y-%m-%d') if customer.last_purchase else 'Never'
            ])

        return response

class LowStockReportView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/reports/low_stock_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get system settings for low stock threshold or use default
        try:
            settings = SystemSettings.objects.first()
            threshold = settings.low_stock_threshold if settings else 10
        except:
            threshold = 10
            
        # Get all products with stock below threshold - using quantity_in_stock instead of current_stock
        low_stock_products = Product.objects.filter(
            quantity_in_stock__lte=models.F('minimum_stock_level')
        ).order_by('quantity_in_stock')
        
        context['low_stock_products'] = low_stock_products
        context['threshold'] = threshold
        return context

# Add these new views for export functionality
class LowStockExportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Create a file-like buffer to receive Excel data
        output = BytesIO()
        
        # Create the Excel workbook and add a worksheet
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Low Stock Products')
        
        # Add styles
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#f8f9fc',
            'border': 1
        })
        
        # Write headers
        headers = ['Product Name', 'SKU', 'Category', 'Current Stock', 'Minimum Level', 'Status']
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Get low stock products
        low_stock_products = Product.objects.filter(
            quantity_in_stock__lte=models.F('minimum_stock_level')
        ).order_by('quantity_in_stock')
        
        # Write data rows
        for row_num, product in enumerate(low_stock_products, 1):
            # Determine status
            if product.quantity_in_stock == 0:
                status = 'Out of Stock'
            else:
                status = 'Low Stock'
                
            worksheet.write(row_num, 0, product.name)
            worksheet.write(row_num, 1, product.sku or '-')
            worksheet.write(row_num, 2, product.category.name if product.category else '-')
            worksheet.write(row_num, 3, product.quantity_in_stock)
            worksheet.write(row_num, 4, product.minimum_stock_level)
            worksheet.write(row_num, 5, status)
        
        # Set column widths
        worksheet.set_column(0, 0, 30)  # Product name
        worksheet.set_column(1, 1, 15)  # SKU
        worksheet.set_column(2, 2, 20)  # Category
        worksheet.set_column(3, 4, 15)  # Stock levels
        worksheet.set_column(5, 5, 15)  # Status
        
        # Close the workbook
        workbook.close()
        
        # Set up the response
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=low_stock_report.xlsx'
        
        return response

class LowStockExportPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Get low stock products
        low_stock_products = Product.objects.filter(
            quantity_in_stock__lte=models.F('minimum_stock_level')
        ).order_by('quantity_in_stock')
        
        # Create a file-like buffer to receive PDF data
        buffer = BytesIO()
        
        # Create the PDF object, using the BytesIO object as its "file."
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = 1  # Center alignment

        # Add title
        elements.append(Paragraph("Low Stock Report", title_style))
        elements.append(Spacer(1, 0.25*inch))
        
        # Add date
        date_style = styles['Normal']
        date_style.alignment = 1  # Center alignment
        elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
        elements.append(Spacer(1, 0.25*inch))
        
        # Create table data
        data = [['Product', 'SKU', 'Category', 'Current Stock', 'Minimum Level', 'Status']]
        
        # Add product data
        for product in low_stock_products:
            if product.quantity_in_stock == 0:
                status = 'Out of Stock'
            else:
                status = 'Low Stock'
                
            data.append([
                product.name,
                product.sku or '-',
                product.category.name if product.category else '-',
                str(product.quantity_in_stock),
                str(product.minimum_stock_level),
                status
            ])
        
        # If no products, add a message
        if len(data) == 1:
            data.append(['No low stock items found.', '', '', '', '', ''])
        
        # Create table
        table = Table(data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1*inch, 1*inch, 1*inch])
        
        # Add style to table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])

        # Apply conditional formatting for status column
        for i in range(1, len(data)):
            if data[i][-1] == 'Out of Stock':
                style.add('TEXTCOLOR', (-1, i), (-1, i), colors.red)
                style.add('FONTNAME', (-1, i), (-1, i), 'Helvetica-Bold')
            elif data[i][-1] == 'Low Stock':
                style.add('TEXTCOLOR', (-1, i), (-1, i), colors.orange)
                style.add('FONTNAME', (-1, i), (-1, i), 'Helvetica-Bold')
        
        table.setStyle(style)
        elements.append(table)
        
        # Add footer
        elements.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'footer',
            parent=styles['Normal'],
            alignment=1,
            fontSize=8,
            textColor=colors.gray
        )
        elements.append(Paragraph("Solar Inventory System - Low Stock Report", footer_style))

        # Build the PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create the HTTP response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="low_stock_report.pdf"'
        response.write(pdf)
        
        return response

# User Profile and Settings Views
class UserProfileView(LoginRequiredMixin, UpdateView):
    form_class = UserProfileForm
    template_name = 'inventory/user/profile.html'
    success_url = reverse_lazy('inventory:user-profile')

    def get_object(self):
        return self.request.user
    
    def form_valid(self, form):
        messages.success(self.request, "Profile updated successfully!")
        return super().form_valid(form)

class SettingsView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/user/settings.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get or create system settings
        settings, created = SystemSettings.objects.get_or_create(
            defaults={
                'company_name': 'Solar Inventory System',
                'company_address': 'Default Address',
                'company_phone': '123-456-7890',
                'company_email': 'info@example.com',
            }
        )
        
        # If form was submitted but had errors, use the form from kwargs
        form = kwargs.get('form', SettingsForm(instance=settings))
        context['form'] = form
        return context
    
    def post(self, request, *args, **kwargs):
        # Get existing settings
        settings, created = SystemSettings.objects.get_or_create()
        
        # Process the form data
        form = SettingsForm(request.POST, request.FILES, instance=settings)
        
        if form.is_valid():
            form.save()
            messages.success(request, "Settings updated successfully!")
            return redirect('inventory:settings')
        else:
            # Print form errors for debugging
            print(f"Form errors: {form.errors}")
            messages.error(request, "There was an error saving your settings. Please check the form.")

        # If form is invalid, re-render the page with the form errors
        return self.render_to_response(self.get_context_data(form=form))

# User Management (Admin only)
class StaffRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff

class UserManagementView(LoginRequiredMixin, StaffRequiredMixin, ListView):
    template_name = 'inventory/admin/user_management.html'
    context_object_name = 'users'
    
    def get_queryset(self):
        User = get_user_model()
        return User.objects.all().order_by('username')
    
    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        User = get_user_model()
        
        if action == 'add_user':
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            first_name = request.POST.get('first_name', '')
            last_name = request.POST.get('last_name', '')
            is_staff = 'is_staff' in request.POST
            is_sales_person = 'is_sales_person' in request.POST
            is_inventory_manager = 'is_inventory_manager' in request.POST
            
            # Check if username already exists
            if User.objects.filter(username=username).exists():
                messages.error(request, f"Username '{username}' already exists.")
                return redirect('inventory:user-management')
            
            # Create new user
            try:
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
                user.is_staff = is_staff
                user.is_sales_person = is_sales_person
                user.is_inventory_manager = is_inventory_manager
                user.save()
                
                messages.success(request, f"User '{username}' created successfully.")
            except Exception as e:
                messages.error(request, f"Error creating user: {str(e)}")
            
        elif action == 'edit_user':
            user_id = request.POST.get('user_id')
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            first_name = request.POST.get('first_name', '')
            last_name = request.POST.get('last_name', '')
            is_staff = 'is_staff' in request.POST
            is_sales_person = 'is_sales_person' in request.POST
            is_inventory_manager = 'is_inventory_manager' in request.POST
            
            try:
                user = User.objects.get(id=user_id)
                
                # Check if username is being changed and already exists
                if user.username != username and User.objects.filter(username=username).exists():
                    messages.error(request, f"Username '{username}' already exists.")
                    return redirect('inventory:user-management')
                
                user.username = username
                user.email = email
                user.first_name = first_name
                user.last_name = last_name
                user.is_staff = is_staff
                user.is_sales_person = is_sales_person
                user.is_inventory_manager = is_inventory_manager
                
                # Only set password if provided
                if password:
                    user.set_password(password)
                
                user.save()
                messages.success(request, f"User '{username}' updated successfully.")
            except User.DoesNotExist:
                messages.error(request, "User not found.")
            except Exception as e:
                messages.error(request, f"Error updating user: {str(e)}")
            
        elif action == 'delete_user':
            user_id = request.POST.get('user_id')
            
            try:
                user = User.objects.get(id=user_id)
                username = user.username
                
                # Prevent deletion of superusers
                if user.is_superuser:
                    messages.error(request, "Cannot delete superuser accounts.")
                    return redirect('inventory:user-management')
                
                user.delete()
                messages.success(request, f"User '{username}' deleted successfully.")
            except User.DoesNotExist:
                messages.error(request, "User not found.")
            except Exception as e:
                messages.error(request, f"Error deleting user: {str(e)}")
        
        return redirect('inventory:user-management')

# Add this view function
@login_required
def customer_create_ajax(request):
    if request.method == 'POST':
        # Create a form instance with the POST data
        form = CustomerForm(request.POST)

        if form.is_valid():
            customer = form.save()
            # Return JSON response with the new customer data
            return JsonResponse({
                'success': True,
                'customer': {
                    'id': customer.id,
                    'name': customer.name,
                }
            })
        else:
            # Return form errors
            return JsonResponse({
                'success': False,
                'error': 'Form validation failed',
                'errors': form.errors.as_json()
            })

    # If not POST, return error
    return JsonResponse({
        'success': False,
        'error': 'Only POST requests are allowed'
    })

def product_search_barcode(request):
    barcode = request.GET.get('barcode', '')
    try:
        product = Product.objects.get(barcode=barcode)
        return JsonResponse({
            'success': True,
            'product': {
                'id': product.id,
                'name': product.name,
                'unit_price': float(product.unit_price),
                'quantity_in_stock': product.quantity_in_stock
            }
        })
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Product not found'
        })

@login_required
def sale_email_invoice(request):
    """Email invoice to customer"""
    if request.method == 'POST':
        sale_id = request.POST.get('sale_id')
        if not sale_id:
            return JsonResponse({'success': False, 'error': _('Sale ID is required')})

        try:
            sale = Sale.objects.get(pk=sale_id)

            # Check if customer has email
            if not sale.customer or not sale.customer.email:
                return JsonResponse({'success': False, 'error': _('Customer has no email address')})
            
            # Generate PDF invoice
            try:
                pdf_content = generate_invoice_pdf(sale)
                
                # Send email with PDF attachment
                subject = f'Invoice #{sale.invoice_number}'
                message = f'Please find attached your invoice #{sale.invoice_number}.'
                from_email = settings.DEFAULT_FROM_EMAIL
                recipient_list = [sale.customer.email]

                email = EmailMessage(subject, message, from_email, recipient_list)
                email.attach(f'invoice_{sale.invoice_number}.pdf', pdf_content, 'application/pdf')
                email.send()

                return JsonResponse({'success': True})
            except Exception as e:
                logger.error(f"Error generating or sending PDF: {str(e)}")
                return JsonResponse({'success': False, 'error': str(e)})
                
        except Sale.DoesNotExist:
            return JsonResponse({'success': False, 'error': _('Sale not found')})
        except Exception as e:
            logger.error(f"Error in sale_email_invoice: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        # For GET requests, redirect to the sale detail page
        sale_id = request.GET.get('sale_id')
        if sale_id:
            messages.info(request, _('Please use the email button on the sale detail page'))
            return redirect('inventory:sale-detail', pk=sale_id)
        else:
            messages.error(request, _('Sale ID is required'))
            return redirect('inventory:sale-list')

def email_proforma(request):
    """Email a proforma invoice to a customer"""
    proforma_id = request.POST.get('proforma_id')

    try:
        proforma = Sale.objects.get(id=proforma_id, is_proforma=True)

        if not proforma.customer or not proforma.customer.email:
            return JsonResponse({'success': False, 'error': 'Customer has no email address'})

        # Generate PDF proforma invoice
        pdf_content = generate_invoice_pdf(proforma, is_proforma=True)

        # Send email with PDF attachment
        subject = f'Proforma Invoice #{proforma.invoice_number}'
        message = f'Please find attached your proforma invoice #{proforma.invoice_number}.'
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [proforma.customer.email]

        email = EmailMessage(subject, message, from_email, recipient_list)
        email.attach(f'proforma_{proforma.invoice_number}.pdf', pdf_content, 'application/pdf')
        email.send()

        return JsonResponse({'success': True})
    except Sale.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proforma invoice not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def invoice_content(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)

    # Calculate any additional fields needed for the invoice
    if not hasattr(sale, 'balance_due'):
        sale.balance_due = sale.total_amount - (sale.amount_paid or 0)

    context = {
        'sale': sale,
    }

    return render(request, 'inventory:invoice_partial.html', context)

@login_required
@require_POST
def hold_sale(request):
    """
    Save a sale as held for later processing
    """
    try:
        data = json.loads(request.body)

        customer_id = data.get('customer_id')
        notes = data.get('notes', '')
        items = data.get('items', [])

        if not items:
            return JsonResponse({
                'success': False,
                'error': _('No items provided')
            })

        with transaction.atomic():
            # Create the held sale
            held_sale = HeldSale(
                user=request.user,
                notes=notes
            )

            # Set customer if provided
            if customer_id:
                try:
                    customer = Customer.objects.get(id=customer_id)
                    held_sale.customer = customer
                except Customer.DoesNotExist:
                    pass
                
            held_sale.save()

            # Add items to the held sale
            total_amount = Decimal('0.00')

            for item_data in items:
                product_id = item_data.get('product_id')
                quantity = Decimal(item_data.get('quantity', 0))
                unit_price = Decimal(item_data.get('unit_price', 0))
                discount = Decimal(item_data.get('discount', 0))

                if not product_id or quantity <= 0:
                    continue

                try:
                    product = Product.objects.get(id=product_id)

                    # Calculate item total
                    item_total = (unit_price * quantity) - discount
                    total_amount += item_total

                    # Create the held sale item
                    HeldSaleItem.objects.create(
                        held_sale=held_sale,
                        product=product,
                        quantity=quantity,
                        unit_price=unit_price,
                        discount_amount=discount,
                        total_price=item_total
                    )
                except Product.DoesNotExist:
                    continue

            # Update the total amount
            held_sale.total_amount = total_amount
            held_sale.save()

            return JsonResponse({
                'success': True,
                'sale_id': held_sale.id
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_GET
def get_held_sales(request):
    """
    Get all held sales for the current user
    """
    held_sales = HeldSale.objects.filter(user=request.user).order_by('-created_at')

    # Annotate with item count
    for sale in held_sales:
        sale.item_count = sale.heldsaleitem_set.count()

    context = {
        'held_sales': held_sales
    }

    html = render_to_string('inventory/sales/held_sales_list.html', context)

    return JsonResponse({
        'success': True,
        'html': html,
        'count': held_sales.count()
    })

@login_required
@require_GET
def get_held_sale(request, sale_id):
    """
    Get a specific held sale
    """
    try:
        held_sale = get_object_or_404(HeldSale, id=sale_id, user=request.user)
        
        # Get all items
        items = []
        for item in held_sale.heldsaleitem_set.all():
            items.append({
                'product_id': item.product.id,
                'product_name': item.product.name,
                'quantity': str(item.quantity),
                'unit_price': str(item.unit_price),
                'discount': str(item.discount_amount),
                'total_price': str(item.total_price)
            })

        # Prepare response data
        sale_data = {
            'id': held_sale.id,
            'customer_id': held_sale.customer.id if held_sale.customer else None,
            'notes': held_sale.notes,
            'total_amount': str(held_sale.total_amount),
            'created_at': held_sale.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'items': items
        }

        # Delete the held sale after loading
        held_sale.delete()

        return JsonResponse({
            'success': True,
            'sale': sale_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_POST
def delete_held_sale(request, sale_id):
    """
    Delete a held sale
    """
    try:
        held_sale = get_object_or_404(HeldSale, id=sale_id, user=request.user)
        held_sale.delete()
        
        return JsonResponse({
            'success': True
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

def proforma_list(request):
    """Return a list of proforma invoices as JSON"""
    proformas = Sale.objects.filter(is_proforma=True)

    proforma_list = []
    for proforma in proformas:
        proforma_list.append({
            'id': proforma.id,
            'invoice_number': proforma.invoice_number,
            'customer_name': proforma.customer.name if proforma.customer else 'Walk-in Customer',
            'date': proforma.date.strftime('%Y-%m-%d'),
            'due_date': proforma.due_date.strftime('%Y-%m-%d') if proforma.due_date else '',
            'total': float(proforma.get_total()),
            'status': proforma.status
        })
    
    return JsonResponse({'proformas': proforma_list})

def proforma_detail(request):
    """Return the HTML content for a proforma invoice"""
    proforma_id = request.GET.get('id')

    try:
        proforma = Sale.objects.get(id=proforma_id, is_proforma=True)
        
        # Render the proforma invoice template
        html = render_to_string('inventory/sales/proforma_invoice_partial.html', {
            'sale': proforma,
            'items': proforma.saleitems_set.all(),
            'currency_symbol': settings.CURRENCY_SYMBOL
        })

        return JsonResponse({'html': html})
    except Sale.DoesNotExist:
        return JsonResponse({'error': 'Proforma invoice not found'}, status=404)

def extract_formset_data(post_data, prefix='items'):
    """Helper function to extract formset data from POST"""
    total_forms = int(post_data.get(f'{prefix}-TOTAL_FORMS', 0))
    items = []
    
    for i in range(total_forms):
        item = {}
        for key in post_data:
            if key.startswith(f'{prefix}-{i}-'):
                field_name = key.replace(f'{prefix}-{i}-', '')
                item[field_name] = post_data[key]
        
        if item:
            items.append(item)
    
    return items

@login_required
@require_POST
def convert_proforma_to_sale(request, pk):
    """Convert a proforma invoice to a regular sale"""
    try:
        proforma = Sale.objects.get(pk=pk, is_proforma=True)
        
        # Extract form data
        payment_method = request.POST.get('payment_method')
        reference_number = request.POST.get('reference_number', '')
        amount_tendered = Decimal(request.POST.get('amount_tendered', 0))
        change_amount = Decimal(request.POST.get('change_amount', 0))
        notes = request.POST.get('notes', '')
        
        # Validate payment method
        if payment_method not in dict(Sale.PAYMENT_METHOD_CHOICES):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': _('Invalid payment method')
                })
            else:
                messages.error(request, _('Invalid payment method'))
                return redirect('inventory:sale-detail', pk=proforma.id)
        
        # Validate reference number for card/mobile payments
        if payment_method in [Sale.CARD, Sale.MOBILE] and not reference_number:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': _('Reference number is required for card/mobile payments')
                })
            else:
                messages.error(request, _('Reference number is required for card/mobile payments'))
                return redirect('inventory:sale-detail', pk=proforma.id)
        
        # Validate amount tendered for cash payments
        if payment_method == Sale.CASH and amount_tendered < proforma.total_amount:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': _('Amount tendered must be at least equal to the total amount')
                })
            else:
                messages.error(request, _('Amount tendered must be at least equal to the total amount'))
                return redirect('inventory:sale-detail', pk=proforma.id)
        
        # Convert the proforma to a sale
        try:
            # Update proforma fields
            proforma.is_proforma = False
            proforma.paid = True
            proforma.payment_method = payment_method
            proforma.reference_number = reference_number
            proforma.amount_tendered = amount_tendered
            proforma.change_amount = change_amount
            
            # Append notes if provided
            if notes:
                if proforma.notes:
                    proforma.notes += f"\n\n{notes}"
                else:
                    proforma.notes = notes
            
            # Generate a new invoice number (from proforma to regular)
            last_sale = Sale.objects.filter(is_proforma=False).order_by('-id').first()
            last_id = last_sale.id if last_sale else 0
            proforma.invoice_number = f"INV-{last_id + 1:06d}"
            
            # Save the changes
            proforma.save()

            # Update inventory
            for item in proforma.items.all():
                product = item.product
                if item.quantity > product.quantity_in_stock:
                    raise ValidationError(_("Not enough stock for {}. Only {} available.").format(
                        product.name, product.quantity_in_stock
                    ))
                
                product.quantity_in_stock -= item.quantity
                product.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'redirect_url': reverse('inventory:sale-detail', kwargs={'pk': proforma.id})
                })
            else:
                messages.success(request, _('Proforma invoice successfully converted to sale'))
                return redirect('inventory:sale-detail', pk=proforma.id)

        except ValidationError as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                })
            else:
                messages.error(request, str(e))
                return redirect('inventory:sale-detail', pk=proforma.id)
    
    except Sale.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': _('Proforma invoice not found')
            })
        else:
            messages.error(request, _('Proforma invoice not found'))
            return redirect('inventory:sales')
    
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
        else:
            messages.error(request, _('Error: {}').format(e))
            return redirect('inventory:sale-detail', pk=pk)
